from PyQt5 import QtWidgets, QtGui, QtCore
from PyQt5.QtWidgets import QApplication, QMainWindow, QTableWidgetItem, QHeaderView
import sys
import random
import string
from openpyxl import Workbook, load_workbook
import os
import pyperclip

# Configure the Excel workbook
file_path = os.path.join(os.path.dirname(__file__), "passwords.xlsx")
if os.path.exists(file_path):
    workbook = load_workbook(file_path)
    sheet = workbook.active
else:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Passwords"
    sheet.append(["Website/App", "Username", "Password"])
    workbook.save(file_path)

# Function to generate passwords
def generate_password(length=12, use_uppercase=True, use_numbers=True, use_symbols=True):
    characters = string.ascii_lowercase
    if use_uppercase:
        characters += string.ascii_uppercase
    if use_numbers:
        characters += string.digits
    if use_symbols:
        characters += string.punctuation
    
    return ''.join(random.choice(characters) for _ in range(length))

class PasswordGeneratorApp(QMainWindow):
    def __init__(self):
        super(PasswordGeneratorApp, self).__init__()
        self.setWindowTitle("Password Generator")
        self.setGeometry(200, 200, 1000, 700)
        self.initUI()

    def initUI(self):
        # Central widget and layout
        central_widget = QtWidgets.QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QtWidgets.QVBoxLayout()
        main_layout.setContentsMargins(20, 20, 20, 20)  # Add padding around the main layout
        central_widget.setLayout(main_layout)

        # Set larger font for better readability
        font = QtGui.QFont()
        font.setPointSize(14)
        self.setFont(font)

        # Input form layout
        form_layout = QtWidgets.QFormLayout()
        main_layout.addLayout(form_layout)
        
        # Labels and Input Fields
        self.website_input = QtWidgets.QLineEdit()
        form_layout.addRow("Website/App:", self.website_input)
        
        self.username_input = QtWidgets.QLineEdit()
        form_layout.addRow("Username:", self.username_input)
        
        self.password_input = QtWidgets.QLineEdit()
        form_layout.addRow("Password:", self.password_input)
        
        # Password Generation Settings
        settings_layout = QtWidgets.QHBoxLayout()
        settings_layout.setSpacing(10)
        self.length_label = QtWidgets.QLabel("Password Length:")
        settings_layout.addWidget(self.length_label)
        
        self.length_input = QtWidgets.QSpinBox()
        self.length_input.setValue(12)
        settings_layout.addWidget(self.length_input)
        
        self.generate_button = QtWidgets.QPushButton("Generate Password")
        self.generate_button.clicked.connect(self.generate_password)
        self.generate_button.setStyleSheet("background-color: #4CAF50; color: white; font-weight: bold; height: 50px; font-size: 18px;")
        settings_layout.addWidget(self.generate_button)
        
        form_layout.addRow(settings_layout)

        self.uppercase_checkbox = QtWidgets.QCheckBox("Include Uppercase Letters")
        self.uppercase_checkbox.setChecked(True)
        form_layout.addRow(self.uppercase_checkbox)
        
        self.numbers_checkbox = QtWidgets.QCheckBox("Include Numbers")
        self.numbers_checkbox.setChecked(True)
        form_layout.addRow(self.numbers_checkbox)
        
        self.symbols_checkbox = QtWidgets.QCheckBox("Include Symbols")
        self.symbols_checkbox.setChecked(True)
        form_layout.addRow(self.symbols_checkbox)

        # Center checkboxes
        checkboxes_layout = QtWidgets.QHBoxLayout()
        checkboxes_layout.addStretch(1)
        checkboxes_layout.addWidget(self.uppercase_checkbox)
        checkboxes_layout.addWidget(self.numbers_checkbox)
        checkboxes_layout.addWidget(self.symbols_checkbox)
        checkboxes_layout.addStretch(1)
        main_layout.addLayout(checkboxes_layout)

        # Buttons layout
        buttons_layout = QtWidgets.QHBoxLayout()
        buttons_layout.setSpacing(20)
        main_layout.addLayout(buttons_layout)

        self.save_button = QtWidgets.QPushButton("Save Password")
        self.save_button.clicked.connect(self.save_password)
        self.save_button.setStyleSheet("background-color: #2196F3; color: white; font-weight: bold; height: 50px; font-size: 18px;")
        buttons_layout.addWidget(self.save_button)
        
        self.show_button = QtWidgets.QPushButton("Show Saved Passwords")
        self.show_button.clicked.connect(self.toggle_passwords)
        self.show_button.setStyleSheet("background-color: #FF9800; color: white; font-weight: bold; height: 50px; font-size: 18px;")
        buttons_layout.addWidget(self.show_button)

        # Table to display saved passwords
        self.passwords_table = QtWidgets.QTableWidget()
        self.passwords_table.setColumnCount(4)
        self.passwords_table.setHorizontalHeaderLabels(["Website/App", "Username", "Password", "Show Password"])
        self.passwords_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.passwords_table.setFont(font)
        main_layout.addWidget(self.passwords_table)
        self.passwords_table.hide()

    def generate_password(self):
        length = self.length_input.value()
        use_uppercase = self.uppercase_checkbox.isChecked()
        use_numbers = self.numbers_checkbox.isChecked()
        use_symbols = self.symbols_checkbox.isChecked()
        
        password = generate_password(length, use_uppercase, use_numbers, use_symbols)
        self.password_input.setText(password)

    def save_password(self):
        website = self.website_input.text()
        username = self.username_input.text()
        password = self.password_input.text()
        
        if not website or not username or not password:
            QtWidgets.QMessageBox.warning(self, "Incomplete Data", "Please fill in all fields.")
            return
        
        sheet.append([website, username, password])
        workbook.save(file_path)
        QtWidgets.QMessageBox.information(self, "Saved", "Password saved successfully!")
        
        self.website_input.clear()
        self.username_input.clear()
        self.password_input.clear()

    def toggle_passwords(self):
        if self.passwords_table.isVisible():
            self.passwords_table.hide()
            self.show_button.setText("Show Saved Passwords")
        else:
            self.display_passwords()
            self.passwords_table.show()
            self.show_button.setText("Hide Saved Passwords")

    def display_passwords(self):
        self.passwords_table.setRowCount(0)
        for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
            self.passwords_table.insertRow(idx)
            self.passwords_table.setItem(idx, 0, QTableWidgetItem(row[0]))
            self.passwords_table.setItem(idx, 1, QTableWidgetItem(row[1]))
            password_item = QTableWidgetItem("**********")
            password_item.setFlags(password_item.flags() & ~QtCore.Qt.ItemIsEditable)
            self.passwords_table.setItem(idx, 2, password_item)
            
            show_button = QtWidgets.QPushButton("Show Password")
            show_button.setCheckable(True)
            show_button.clicked.connect(lambda state, row=idx: self.toggle_password_visibility(row, state))
            self.passwords_table.setCellWidget(idx, 3, show_button)

            # Add copy functionality
            password_item.setToolTip("Click to copy password")
            self.passwords_table.item(idx, 2).setFlags(password_item.flags() | QtCore.Qt.ItemIsSelectable)
            self.passwords_table.item(idx, 2).setTextAlignment(QtCore.Qt.AlignCenter)

        # Connect copy functionality to single selection
        self.passwords_table.itemClicked.connect(self.copy_password_to_clipboard)

    def toggle_password_visibility(self, row, state):
        if state:
            self.passwords_table.item(row, 2).setText(sheet.cell(row=row + 2, column=3).value)
            self.passwords_table.cellWidget(row, 3).setText("Hide Password")
        else:
            self.passwords_table.item(row, 2).setText("**********")
            self.passwords_table.cellWidget(row, 3).setText("Show Password")

    def copy_password_to_clipboard(self, item):
        if item.column() == 2:
            row = item.row()
            actual_password = sheet.cell(row=row + 2, column=3).value
            pyperclip.copy(actual_password)
            QtWidgets.QMessageBox.information(self, "Copied", "Password copied to clipboard!")
            self.passwords_table.blockSignals(True)  # Block signals to prevent multiple pop-ups
            self.passwords_table.clearSelection()
            self.passwords_table.blockSignals(False)  # Re-enable signals

if __name__ == "__main__":
    app = QApplication(sys.argv)
    main_window = PasswordGeneratorApp()
    main_window.show()
    sys.exit(app.exec_())
